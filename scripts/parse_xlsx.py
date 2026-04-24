# -*- coding: utf-8 -*-
"""xlsx / csv question-bank adapter.

Expected columns (auto-detected, Chinese / English):
  chapter   章节 / 章 / Chapter
  type      题型 / Type             (single|multi|tf|fill|short / 单选/多选/判断/填空/简答)
  question  题目 / 题干 / Question
  A B C D E F (or 选项A 等)
  answer    答案 / Answer
  explanation 解析 / Explanation     (optional)

Outputs questions.json conforming to SCHEMA.md v2.
"""
import argparse, json, os, re, sys, io
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from common import (
    file_sha256, get_version, utc_now_iso, question_hash,
)


HEADER_ALIASES = {
    "chapter":     ["chapter", "章节", "章", "section", "单元"],
    "type":        ["type", "题型", "类型"],
    "question":    ["question", "题目", "题干", "stem"],
    "answer":      ["answer", "答案", "ans", "正确答案", "参考答案"],
    "explanation": ["explanation", "解析", "答案解析", "详解", "考点"],
}
OPT_ALIASES = {
    L: [L, L.lower(), f"选项{L}", f"option {L}".lower(), f"option_{L.lower()}"]
    for L in "ABCDEF"
}
TYPE_NORM = {
    "单选": "single", "单选题": "single", "single": "single",
    "多选": "multi",  "多选题": "multi",  "multi": "multi", "multiple": "multi",
    "判断": "tf",    "判断题": "tf",    "tf": "tf", "true/false": "tf", "true_false": "tf",
    "填空": "fill",  "填空题": "fill",  "fill": "fill", "fill_blank": "fill",
    "简答": "short", "简答题": "short", "short": "short", "subjective": "short",
}


def normalize_header(h: str) -> str:
    return (h or "").strip().lower().replace(" ", "")


def find_col(headers_lower, candidates):
    for c in candidates:
        cl = c.strip().lower().replace(" ", "")
        for i, h in enumerate(headers_lower):
            if h == cl:
                return i
    return -1


def normalize_answer(raw, qtype):
    if qtype in ("fill", "short"):
        if raw is None:
            return []
        return [p.strip() for p in re.split(r"[;；/|｜,，]\s*", str(raw).strip()) if p.strip()]
    s = str(raw or "").strip().upper().replace(" ", "")
    if qtype == "tf":
        if s in ("A", "B"):
            return s
        if s in ("对", "正确", "是", "T", "TRUE", "Y", "YES", "√", "✓", "1"):
            return "A"
        if s in ("错", "错误", "否", "F", "FALSE", "N", "NO", "×", "✗", "0"):
            return "B"
        return s[:1] if s and s[0] in "AB" else ""
    return "".join(c for c in s if c in "ABCDEF")


def normalize_type(raw):
    s = str(raw or "").strip().lower().replace(" ", "")
    return TYPE_NORM.get(s, "single")


def iter_rows(path: str):
    """Yield row dicts from xlsx or csv. First row is treated as header."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        import csv
        # Try utf-8 first, fall back to gbk
        for enc in ("utf-8-sig", "utf-8", "gbk"):
            try:
                with open(path, encoding=enc, newline="") as f:
                    reader = csv.reader(f)
                    headers = next(reader, [])
                    yield headers, [row for row in reader]
                return
            except UnicodeDecodeError:
                continue
        sys.exit("ERROR: cannot decode CSV (tried utf-8/gbk)")
    elif ext in (".xlsx", ".xls"):
        try:
            import openpyxl
        except ImportError:
            sys.exit("ERROR: pip install openpyxl")
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            sys.exit("ERROR: empty workbook")
        headers = [str(c) if c is not None else "" for c in rows[0]]
        body = [[c for c in r] for r in rows[1:]]
        yield headers, body
    else:
        sys.exit(f"ERROR: unsupported {ext}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--source-id", default="s1")
    ap.add_argument("--strict", action="store_true")
    ap.add_argument("--default-type", default="single",
                    help="When 'type' column is missing, fallback type")
    args = ap.parse_args()

    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

    headers, body = next(iter_rows(args.input))
    h_lower = [normalize_header(h) for h in headers]
    cols = {k: find_col(h_lower, v) for k, v in HEADER_ALIASES.items()}
    opt_cols = {L: find_col(h_lower, OPT_ALIASES[L]) for L in "ABCDEF"}

    if cols["question"] < 0:
        sys.exit(
            f"ERROR: required column 'question/题目' not found. Headers seen: {headers}"
        )
    if cols["answer"] < 0:
        sys.exit(
            f"ERROR: required column 'answer/答案' not found. Headers seen: {headers}"
        )

    chapters_seen = []
    questions = []
    issues = 0
    for ridx, row in enumerate(body, start=2):
        if not any(c not in (None, "") for c in row):
            continue
        def cell(idx):
            return row[idx] if 0 <= idx < len(row) else None
        chap = str(cell(cols["chapter"]) or "未分类").strip()
        if chap not in chapters_seen:
            chapters_seen.append(chap)
        qtype_raw = cell(cols["type"])
        qtype = normalize_type(qtype_raw) if qtype_raw is not None else args.default_type
        question = str(cell(cols["question"]) or "").strip()
        if not question:
            continue
        options = {}
        for L in "ABCDEF":
            v = cell(opt_cols[L])
            if v is not None and str(v).strip() != "":
                options[L] = str(v).strip()
        if qtype == "tf" and not options:
            options = {"A": "对", "B": "错"}
        ans = normalize_answer(cell(cols["answer"]), qtype)
        explanation = ""
        if cols["explanation"] >= 0:
            explanation = str(cell(cols["explanation"]) or "").strip()
        q = {
            "id": len(questions) + 1,
            "source_id": args.source_id,
            "chapter": chap,
            "type": qtype,
            "question": question,
            "options": options,
            "answer": ans,
            "explanation": explanation,
        }
        # Validation
        if qtype in ("single", "multi"):
            for L in (ans if isinstance(ans, str) else ""):
                if L not in options:
                    issues += 1
        elif qtype == "tf" and ans not in ("A", "B"):
            issues += 1
        q["qhash"] = question_hash(q)
        questions.append(q)

    by_type = {}
    for q in questions:
        by_type[q["type"]] = by_type.get(q["type"], 0) + 1
    print(f"Total questions: {len(questions)}")
    for t, n in sorted(by_type.items()):
        print(f"  {t}: {n}")
    print(f"Chapters: {len(chapters_seen)}")
    print(f"Validation issues: {issues}")

    payload = {
        "version": get_version(),
        "generator": "quiz-gen",
        "created_at": utc_now_iso(),
        "sources": [{
            "id": args.source_id,
            "path": str(Path(args.input).resolve()),
            "sha256": file_sha256(args.input),
        }],
        "chapters": chapters_seen,
        "questions": questions,
    }
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"Written: {out}")

    if args.strict and issues:
        sys.exit(2)


if __name__ == "__main__":
    main()
